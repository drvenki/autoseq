import logging

from openpyxl import load_workbook
from clinseq_barcode import *


def extract_clinseq_barcodes(input_filename):
    """
    Extrat clinseq barcodes from the specified input file:

    :param input_filename: Either a .txt listing clinseq barcodes one per line,
    or a .xlsx order form file containing the barcodes.

    :return: A list of validated dash-delimite clinseq barcodes.
    """

    toks = input_filename.split(".")
    if len(toks) < 1:
        raise ValueError("Invalid clinseq barcodes input filename: " + input_filename)

    if toks[-1] == "txt":
        return [line.strip() for line in open(input_filename)
                if clinseq_barcode_is_valid(line.strip())]
    elif toks[-1] == "xlsx":
        return parse_orderform(input_filename)
    else:
        raise ValueError("Invalid clinseq barcodes file type: " + input_filename)


def parse_orderform_block(block_of_values):
    """
    Extract clinseq barcodes from the given list of order form fields. Looks
    in the entries between <SAMPLE ENTRIES> and </SAMPLE ENTRIES> for clinseq barcodes.

    :param block_of_values: List of fields from which to extract clinseq barcodes
    :return: List of validated clinseq barcodes
    """

    validated_clinseq_barcodes = []
    clinseq_barcodes_section = False
    for cell_value in block_of_values:
        if cell_value == "</SAMPLE ENTRIES>":
            clinseq_barcodes_section = False
        if clinseq_barcodes_section:
            if not clinseq_barcode_is_valid(cell_value):
                raise ValueError("Invalid clinseq barcode: " + cell_value)
            validated_clinseq_barcodes.append(cell_value)
        if cell_value == "<SAMPLE ENTRIES>":
            clinseq_barcodes_section = True
        else:
            logging.debug("Ignoring field from order form: " + cell_value)

    return validated_clinseq_barcodes


def parse_orderform_worksheet(order_form_worksheet):
    """
    Extract clinseq barcodes from the given order form excel spreadsheet worksheet.

    :param order_form_worksheet: An openpyxl worksheet.
    :return: List of clinseq barcodes extracted from the worksheet.
    """

    first_column_vals = [row[0].value for row in list(order_form_worksheet.iter_rows()) if
                         row[0].value is not None]

    return parse_orderform_block(first_column_vals)


def parse_orderform(order_form_filename):
    """
    Extract clinseq barcodes from the specified order form.

    :param order_form_filename: An excel spreadsheet filename.
    :return: List of clinseq barcodes extracted from the order form.
    """

    workbook = load_workbook(order_form_filename)
    return parse_orderform_worksheet(workbook.worksheets[0])
