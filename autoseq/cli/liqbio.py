import json
import logging
import os
import sys
import time

import click
from openpyxl import load_workbook
from autoseq.pipeline.liqbio import LiqBioPipeline
from autoseq.util.library import get_libdict
from autoseq.util.path import mkdir, stripsuffix


@click.command()
@click.option('--libdir', default="/tmp", help="directory to search for libraries")
@click.argument('sample', type=click.File('r'))
@click.pass_context
def liqbio(ctx, libdir, sample):
    logging.info("Running Liquid Biopsy pipeline")
    logging.info("Sample is {}".format(sample))

    logging.debug("Reading sample config from {}".format(sample))
    sampledata = json.load(sample)

    if ctx.obj['jobdb']:
        mkdir(os.path.dirname(ctx.obj['jobdb']))

    ctx.obj['pipeline'] = LiqBioPipeline(sampledata=sampledata, refdata=ctx.obj['refdata'],
                                         outdir=ctx.obj['outdir'], libdir=libdir, maxcores=ctx.obj['cores'],
                                         debug=ctx.obj['debug'], runner=ctx.obj['runner'],
                                         jobdb=ctx.obj['jobdb'], dot_file=ctx.obj['dot_file'])

    # start main analysis
    ctx.obj['pipeline'].start()
    #
    logging.info("Waiting for pipeline to finish.")
    while ctx.obj['pipeline'].is_alive():
        logging.debug("Waiting for LiqBioPipeline")
        time.sleep(5)

    # # return_code from run_pipeline() will be != 0 if the pipeline fails
    sys.exit(ctx.obj['pipeline'].exitcode)


@click.command()
@click.option('--outdir', required=True, help="directory to write config files")
@click.argument('orderform', type=str)
@click.pass_context
def liqbio_prepare(ctx, outdir, orderform):
    logging.info("Creating config files from excel orderform")
    libs = parse_orderform(orderform)
    sdids = set([lib['sdid'] for lib in libs])
    for sdid in sdids:
        logging.debug("Creating config file for SDID {}".format(sdid))
        panel_t_lib = [lib['library_id'] for lib in libs if
                       lib['sdid'] == sdid and lib['type'] == "T" and lib['capture_id'] != "WGS"]
        panel_n_lib = [lib['library_id'] for lib in libs if
                       lib['sdid'] == sdid and lib['type'] == "N" and lib['capture_id'] != "WGS"]
        panel_p_libs = [lib['library_id'] for lib in libs if
                        lib['sdid'] == sdid and lib['type'] == "P" and lib['capture_id'] != "WGS"]

        wgs_t_lib = [lib['library_id'] for lib in libs if
                     lib['sdid'] == sdid and lib['type'] == "T" and lib['capture_id'] == "WGS"]
        wgs_n_lib = [lib['library_id'] for lib in libs if
                     lib['sdid'] == sdid and lib['type'] == "N" and lib['capture_id'] == "WGS"]
        wgs_p_libs = [lib['library_id'] for lib in libs if
                      lib['sdid'] == sdid and lib['type'] == "P" and lib['capture_id'] == "WGS"]

        def fix_lib(lib):
            """lib is a vector of libraries.
            If it has no contents, return None
            If it has a single element, return it
            If it has more than one element, raise error
            """
            if lib == []:
                return None
            if len(lib) == 1:
                return lib[0]
            if len(lib) > 1:
                raise ValueError("Too many libs for SDID {}. Expected 1, got {}".format(sdid, lib))

        panel_t_lib = fix_lib(panel_t_lib)
        panel_n_lib = fix_lib(panel_n_lib)
        wgs_t_lib = fix_lib(wgs_t_lib)
        wgs_n_lib = fix_lib(wgs_n_lib)

        d = {'sdid': sdid,
             'panel': {
                 'T': panel_t_lib,
                 'N': panel_n_lib,
                 'P': panel_p_libs
             },
             'wgs': {
                 'T': wgs_t_lib,
                 'N': wgs_n_lib,
                 'P': wgs_p_libs
             }
             }

        fn = "{}/{}.json".format(outdir, sdid)
        with open(fn, 'w') as f:
            json.dump(d, f, sort_keys=True, indent=4)


def parse_orderform(xlsx):
    orderform_name = stripsuffix(os.path.basename(xlsx), ".xlsx")
    workbook = load_workbook(xlsx)
    worksheet = workbook.worksheets[0]
    first_idx = None
    last_idx = None

    libraries = []

    for idx in range(1, 1000):
        cell_value = worksheet.cell(column=1, row=idx).value
        if cell_value == "<SAMPLE ENTRIES>":
            first_idx = idx

        # no need to parse after end tag
        if cell_value == "</SAMPLE ENTRIES>":
            break

        if not cell_value:
            continue

        if first_idx is not None and idx > first_idx:
            library_name = str(worksheet.cell(column=1, row=idx).value)

            lib = get_libdict(library_name)
            logging.debug("Parsed library {}".format(lib))
            if lib['type'] not in ['T', 'N', 'P']:
                logging.warning(
                    "Unexpected library type detected: {} for library {}".format(lib['type'], lib['library_id']))
            libraries.append(lib)

    return libraries
